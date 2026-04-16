"""
PRODUCTION INVENTORY SERVER — Standalone
=========================================
Independent inventory tracking from Style Ledger + ATS committed/allocated.
Completely ignores warehouse stock data.

Deploy as its own Render web service.
Requires: DROPBOX_APP_KEY, DROPBOX_APP_SECRET, DROPBOX_REFRESH_TOKEN,
          AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY, S3_BUCKET

Data flow:
  - Hourly: pulls Style Ledger from Dropbox, diffs against S3 snapshot
  - Items on ledger = "incoming"
  - Items that disappear from ledger (3 consecutive checks) = "in_stock"
  - Committed/Allocated pulled from ATS sheet (Dropbox) for deductions
  - ATS Archive folder cross-checks arrival accuracy
  - Frontend gets data in same shape as the main inventory platform
"""

try:
    from gevent import monkey
    monkey.patch_all()
except ImportError:
    pass

import os
import re
import json
import time
import threading
from datetime import datetime, timedelta
from io import BytesIO

import boto3
from botocore.exceptions import ClientError
from flask import Flask, request, jsonify
from flask_cors import CORS
import requests as http_requests
import openpyxl

# ══════════════════════════════════════════════════
#  APP SETUP
# ══════════════════════════════════════════════════

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*", "methods": ["GET", "POST", "OPTIONS"], "allow_headers": ["Content-Type"]}})

# ══════════════════════════════════════════════════
#  CONFIG — Environment Variables
# ══════════════════════════════════════════════════

# AWS / S3
AWS_REGION = os.environ.get('AWS_REGION', 'us-east-2')
S3_BUCKET = os.environ.get('S3_BUCKET', 'nauticaslimfit')
S3_SNAPSHOT_KEY = os.environ.get('S3_SNAPSHOT_KEY', 'production-inventory/snapshot.json')
S3_HISTORY_PREFIX = 'production-inventory/history/'
S3_SHIPPED_KEY = os.environ.get('S3_SHIPPED_KEY', 'production-inventory/shipped_ledger.json')

# Dropbox OAuth
DROPBOX_APP_KEY = os.environ.get('DROPBOX_APP_KEY', '')
DROPBOX_APP_SECRET = os.environ.get('DROPBOX_APP_SECRET', '')
DROPBOX_REFRESH_TOKEN = os.environ.get('DROPBOX_REFRESH_TOKEN', '')

# Dropbox file paths
DROPBOX_LEDGER_FOLDER = os.environ.get('DROPBOX_LEDGER_FOLDER',
    '/Versa Share Files/David - Dropbox/Style Ledger')
DROPBOX_ATS_PATH = os.environ.get('DROPBOX_ATS_PATH',
    '/Versa Share Files/Hourly ATS/Inventory_ATS.xlsx')
DROPBOX_ARCHIVE_PATH = os.environ.get('DROPBOX_ARCHIVE_PATH',
    '/Versa Share Files/Hourly ATS/ATS_Archive')

# Open Orders API (external — used as-is)
ORDERS_API_URL = os.environ.get('ORDERS_API_URL', 'https://open-orders-api.onrender.com')

# Main inventory API (for proxying production data, overrides, etc.)
MAIN_API_URL = os.environ.get('MAIN_API_URL', 'https://versa-inventory-api.onrender.com')

# Diff engine settings
MISS_THRESHOLD = int(os.environ.get('MISS_THRESHOLD', '3'))
SYNC_INTERVAL = int(os.environ.get('SYNC_INTERVAL', '3600'))  # 1 hour
ARCHIVE_COMPARE_HOURS = int(os.environ.get('ARCHIVE_COMPARE_HOURS', '24'))

# Brand code mapping: Style Ledger → Platform brand name
BRAND_CODE_MAP = {
    'US': 'USPA', 'GB': 'BEENE', 'VC': 'VINCE', 'VD': 'VD',
    'BEN': 'BEN', 'CHAPS': 'CHAPS', 'DKNY': 'DKNY', 'EB': 'EB',
    'JNY': 'JNY', 'KLP': 'KLP', 'LUCKY': 'LUCKY', 'NAUTICA': 'NAUTICA',
    'NE': 'NE', 'RB': 'RB', 'RG': 'RG', 'SHAQ': 'SHAQ',
}

BRAND_FULL_NAMES = {
    'NAUTICA': 'Nautica', 'DKNY': 'DKNY', 'EB': 'Eddie Bauer',
    'VINCE': 'Vince Camuto', 'BEN': 'Ben Sherman', 'USPA': 'U.S. Polo Assn.',
    'CHAPS': 'Chaps', 'LUCKY': 'Lucky Brand', 'JNY': 'Jones New York',
    'BEENE': 'Geoffrey Beene', 'SHAQ': "Shaquille O'Neal", 'VD': 'Von Dutch',
    'KLP': 'Karl Lagerfeld Paris', 'NE': 'Neiman Marcus', 'RB': 'Reebok',
    'RG': 'Robert Graham',
}

# ══════════════════════════════════════════════════
#  DROPBOX AUTH — Auto-refreshing OAuth tokens
# ══════════════════════════════════════════════════

_dbx_token = ''
_dbx_token_expires = 0

def get_dropbox_token():
    global _dbx_token, _dbx_token_expires
    if not DROPBOX_REFRESH_TOKEN or not DROPBOX_APP_KEY:
        return ''
    if time.time() < _dbx_token_expires - 300:
        return _dbx_token
    try:
        resp = http_requests.post('https://api.dropbox.com/oauth2/token', data={
            'grant_type': 'refresh_token',
            'refresh_token': DROPBOX_REFRESH_TOKEN,
            'client_id': DROPBOX_APP_KEY,
            'client_secret': DROPBOX_APP_SECRET,
        }, timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            _dbx_token = data['access_token']
            _dbx_token_expires = time.time() + data.get('expires_in', 14400)
            print(f"[Dropbox] ✓ Token refreshed", flush=True)
            return _dbx_token
        print(f"[Dropbox] ✗ Refresh failed: {resp.status_code}", flush=True)
    except Exception as e:
        print(f"[Dropbox] ✗ Refresh error: {e}", flush=True)
    return _dbx_token  # Return stale token as fallback

# ══════════════════════════════════════════════════
#  S3 CLIENT
# ══════════════════════════════════════════════════

_s3_client = None

def get_s3():
    global _s3_client
    if _s3_client is None:
        _s3_client = boto3.client('s3', region_name=AWS_REGION)
    return _s3_client

# ══════════════════════════════════════════════════
#  DROPBOX FILE FETCHERS
# ══════════════════════════════════════════════════

def _dbx_headers():
    return {'Authorization': f'Bearer {get_dropbox_token()}', 'Content-Type': 'application/json'}

def _dbx_download(path):
    """Download a file from Dropbox. Returns bytes or None."""
    token = get_dropbox_token()
    if not token:
        return None
    resp = http_requests.post(
        'https://content.dropboxapi.com/2/files/download',
        headers={'Authorization': f'Bearer {token}', 'Dropbox-API-Arg': json.dumps({'path': path})},
        timeout=60
    )
    if resp.status_code == 401:
        global _dbx_token_expires
        _dbx_token_expires = 0
        token = get_dropbox_token()
        resp = http_requests.post(
            'https://content.dropboxapi.com/2/files/download',
            headers={'Authorization': f'Bearer {token}', 'Dropbox-API-Arg': json.dumps({'path': path})},
            timeout=60
        )
    if resp.status_code != 200:
        print(f"  ⚠ Dropbox download failed ({resp.status_code}): {path}")
        return None
    return resp.content


def fetch_style_ledger():
    """Fetch and parse the Style Ledger from Dropbox. Returns list of dicts."""
    token = get_dropbox_token()
    if not token:
        print("  ⚠ No Dropbox token for style ledger")
        return []
    try:
        resp = http_requests.post(
            'https://api.dropboxapi.com/2/files/list_folder',
            headers=_dbx_headers(),
            json={'path': DROPBOX_LEDGER_FOLDER, 'recursive': False, 'limit': 50},
            timeout=20
        )
        if resp.status_code != 200:
            print(f"  ⚠ Could not list ledger folder: {resp.status_code}")
            return []

        entries = resp.json().get('entries', [])
        xlsx_files = [e for e in entries if e.get('.tag') == 'file' and e['name'].lower().endswith('.xlsx')]
        if not xlsx_files:
            print(f"  ⚠ No .xlsx in {DROPBOX_LEDGER_FOLDER}")
            return []

        xlsx_files.sort(key=lambda e: (0 if 'ledger' in e['name'].lower() else 1, e['name']))
        chosen = xlsx_files[0]
        print(f"  📋 Downloading style ledger: {chosen['name']}")

        data = _dbx_download(chosen['path_display'])
        if not data:
            return []

        wb = openpyxl.load_workbook(BytesIO(data), read_only=False, data_only=True)
        ws = wb[wb.sheetnames[0]]
        results = []
        for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
            style = str(row[2] or '').strip().upper()
            if not style:
                continue
            etd = None
            if row[5]:
                if isinstance(row[5], datetime):
                    etd = row[5].strftime('%Y-%m-%d')
                else:
                    etd = str(row[5])
            try:
                units = int(row[3] or 0)
            except (ValueError, TypeError):
                units = 0
            results.append({
                'production': str(row[0] or '').strip(),
                'poName': str(row[1] or '').strip(),
                'style': style,
                'units': units,
                'brand': str(row[4] or '').strip().upper(),
                'etd': etd,
            })
        wb.close()
        print(f"  ✓ Parsed {len(results)} production rows")
        return results
    except Exception as e:
        print(f"  ⚠ Style ledger error: {e}")
        return []


def fetch_ats_deductions():
    """Fetch ATS sheet from Dropbox. Returns list of {sku, brand, committed, allocated, incoming}."""
    data = _dbx_download(DROPBOX_ATS_PATH)
    if not data:
        return []
    try:
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        items = []
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row[0]:
                continue
            items.append({
                'sku': str(row[0]).strip().upper(),
                'brand': str(row[1] or '').strip().upper(),
                'incoming': int(row[5] or 0) if len(row) > 5 else 0,
                'committed': int(row[10] or 0) if len(row) > 10 else 0,
                'allocated': int(row[11] or 0) if len(row) > 11 else 0,
            })
        wb.close()
        print(f"  ✓ Parsed {len(items)} ATS rows for deductions")
        return items
    except Exception as e:
        print(f"  ⚠ ATS parse error: {e}")
        return []

# ══════════════════════════════════════════════════
#  ATS ARCHIVE — Cross-check arrivals
# ══════════════════════════════════════════════════

_archive_index = {'files': [], 'last_refresh': 0}
_archive_lock = threading.Lock()

def refresh_archive_index():
    """List ATS_Archive folder, parse dates from filenames, cache for 6 hours."""
    with _archive_lock:
        if _archive_index['files'] and (time.time() - _archive_index['last_refresh'] < 21600):
            return _archive_index['files']

    token = get_dropbox_token()
    if not token:
        return []

    try:
        all_entries = []
        headers = _dbx_headers()
        resp = http_requests.post(
            'https://api.dropboxapi.com/2/files/list_folder',
            headers=headers,
            json={'path': DROPBOX_ARCHIVE_PATH, 'recursive': False, 'limit': 2000},
            timeout=30
        )
        if resp.status_code != 200:
            return []
        data = resp.json()
        all_entries.extend(data.get('entries', []))
        # Paginate
        while data.get('has_more') and data.get('cursor'):
            resp = http_requests.post(
                'https://api.dropboxapi.com/2/files/list_folder/continue',
                headers=headers, json={'cursor': data['cursor']}, timeout=30
            )
            if resp.status_code != 200:
                break
            data = resp.json()
            all_entries.extend(data.get('entries', []))

        files = []
        for e in all_entries:
            if e.get('.tag') != 'file' or not e['name'].lower().endswith('.xlsx'):
                continue
            match = re.search(r'(\d{4}-\d{2}-\d{2})_(\d{6})', e['name'])
            if match:
                try:
                    dt = datetime.strptime(f"{match.group(1)}_{match.group(2)}", '%Y-%m-%d_%H%M%S')
                    files.append({'name': e['name'], 'path': e.get('path_display', ''), 'dt': dt})
                except ValueError:
                    pass

        files.sort(key=lambda f: f['dt'], reverse=True)
        with _archive_lock:
            _archive_index['files'] = files
            _archive_index['last_refresh'] = time.time()
        print(f"  📂 Archive index: {len(files)} files")
        return files
    except Exception as e:
        print(f"  ⚠ Archive index error: {e}")
        return []


def fetch_archive_ats(file_path):
    """Download an archive ATS file, return {sku: {incoming, warehouse}}."""
    data = _dbx_download(file_path)
    if not data:
        return None
    try:
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        result = {}
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row[0]:
                continue
            sku = str(row[0]).strip().upper()
            incoming = int(row[5] or 0) if len(row) > 5 else 0
            warehouse = sum(int(row[i] or 0) for i in range(6, 10) if len(row) > i)
            result[sku] = {'incoming': incoming, 'warehouse': warehouse}
        wb.close()
        return result
    except Exception as e:
        print(f"  ⚠ Archive parse error: {e}")
        return None

# ══════════════════════════════════════════════════
#  S3 SNAPSHOT STORE
# ══════════════════════════════════════════════════

def load_snapshot():
    try:
        resp = get_s3().get_object(Bucket=S3_BUCKET, Key=S3_SNAPSHOT_KEY)
        return json.loads(resp['Body'].read().decode('utf-8'))
    except Exception as e:
        if 'NoSuchKey' in str(e):
            print("  📋 No snapshot found — starting fresh")
            return {'last_check': None, 'orders': {}}
        raise

def save_snapshot(snapshot):
    snapshot['last_check'] = datetime.utcnow().isoformat() + 'Z'
    get_s3().put_object(
        Bucket=S3_BUCKET, Key=S3_SNAPSHOT_KEY,
        Body=json.dumps(snapshot).encode('utf-8'),
        ContentType='application/json'
    )


# ══════════════════════════════════════════════════
#  SHIPPED LEDGER — Tracks cumulative fulfillment
#
#  Solves the "sold-out reappearance" problem:
#  When committed drops (order fulfilled + closed), we record the delta
#  as "shipped_out" permanently — so ATS doesn't rebound when the
#  order falls off the ATS sheet.
#
#  Structure: { "SKU": { shipped_out: N, peak_committed: M, history: [...] } }
# ══════════════════════════════════════════════════

def load_shipped_ledger():
    """Load the shipped ledger from S3. Starts empty if doesn't exist."""
    try:
        resp = get_s3().get_object(Bucket=S3_BUCKET, Key=S3_SHIPPED_KEY)
        return json.loads(resp['Body'].read().decode('utf-8'))
    except Exception as e:
        if 'NoSuchKey' in str(e):
            print("  📋 No shipped ledger found — starting fresh")
            return {'last_update': None, 'skus': {}}
        raise


def save_shipped_ledger(ledger):
    ledger['last_update'] = datetime.utcnow().isoformat() + 'Z'
    get_s3().put_object(
        Bucket=S3_BUCKET, Key=S3_SHIPPED_KEY,
        Body=json.dumps(ledger).encode('utf-8'),
        ContentType='application/json'
    )


def update_shipped_ledger(ats_items, prod_arrived_by_sku):
    """
    Track cumulative deductions per SKU.
    
    Logic:
      - For each SKU, track the peak (highest) committed value we've ever seen
      - When committed drops below peak, the delta is assumed to be shipped out
      - shipped_out only grows, never shrinks
      - Capped at the total arrived production units (can't ship more than we have)
    
    Args:
        ats_items: List of ATS dicts with sku, committed, allocated
        prod_arrived_by_sku: {sku: total_arrived_units} from snapshot
    
    Returns:
        Updated ledger dict
    """
    ledger = load_shipped_ledger()
    now = datetime.utcnow().isoformat() + 'Z'
    updated_count = 0

    # Build quick lookup of current ATS state
    ats_lookup = {i['sku']: i for i in ats_items}

    # Process every SKU we have arrived production for
    for sku, arrived in prod_arrived_by_sku.items():
        if arrived <= 0:
            continue

        ats = ats_lookup.get(sku, {})
        current_committed = abs(ats.get('committed', 0) or 0)
        current_allocated = abs(ats.get('allocated', 0) or 0)

        # Get or initialize this SKU's shipped record
        rec = ledger['skus'].get(sku, {
            'shipped_out': 0,
            'peak_committed': 0,
            'peak_allocated': 0,
            'history': [],
            'first_tracked': now,
        })

        prev_peak_committed = rec.get('peak_committed', 0)
        prev_peak_allocated = rec.get('peak_allocated', 0)

        # Track peak committed + allocated (highest ever seen)
        new_peak_committed = max(prev_peak_committed, current_committed)
        new_peak_allocated = max(prev_peak_allocated, current_allocated)

        # When committed drops from peak, the drop is shipped
        # (peak - current) = amount that has left since peak
        # But we only bump shipped_out by the NEW drop since last check
        prev_drop_committed = prev_peak_committed - (rec.get('prev_committed', prev_peak_committed))
        new_drop_committed = new_peak_committed - current_committed
        committed_shipped_delta = max(0, new_drop_committed - prev_drop_committed)

        prev_drop_allocated = prev_peak_allocated - (rec.get('prev_allocated', prev_peak_allocated))
        new_drop_allocated = new_peak_allocated - current_allocated
        allocated_shipped_delta = max(0, new_drop_allocated - prev_drop_allocated)

        total_new_shipped = committed_shipped_delta + allocated_shipped_delta

        # Cap shipped_out at arrived (can't ship more than we received)
        if total_new_shipped > 0:
            new_shipped_total = min(arrived, rec['shipped_out'] + total_new_shipped)
            if new_shipped_total > rec['shipped_out']:
                delta_applied = new_shipped_total - rec['shipped_out']
                rec['shipped_out'] = new_shipped_total
                rec['history'].append({
                    'date': now,
                    'shipped': delta_applied,
                    'from_committed': committed_shipped_delta,
                    'from_allocated': allocated_shipped_delta,
                })
                # Keep history reasonable
                if len(rec['history']) > 50:
                    rec['history'] = rec['history'][-50:]
                updated_count += 1
                print(f"  📤 {sku} shipped +{delta_applied} (total shipped: {new_shipped_total}/{arrived})")

        # Update tracking fields
        rec['peak_committed'] = new_peak_committed
        rec['peak_allocated'] = new_peak_allocated
        rec['prev_committed'] = current_committed
        rec['prev_allocated'] = current_allocated
        rec['last_seen'] = now
        rec['arrived_total'] = arrived

        ledger['skus'][sku] = rec

    if updated_count > 0:
        save_shipped_ledger(ledger)
        print(f"  ✓ Shipped ledger updated: {updated_count} SKUs recorded new shipments")
    else:
        # Still save to keep prev_committed/prev_allocated fresh
        save_shipped_ledger(ledger)

    return ledger


# ══════════════════════════════════════════════════
#  DIFF ENGINE
# ══════════════════════════════════════════════════

def run_diff(ledger_rows, archive_data=None, current_ats=None):
    """
    Compare current style ledger against S3 snapshot.
    Returns (changes_dict, snapshot).
    """
    snapshot = load_snapshot()
    now = datetime.utcnow().isoformat() + 'Z'

    current_keys = set()
    ledger_map = {}
    # Track occurrences of same production+style for split shipments (different ETDs)
    # Key format: production::style::etd — unique per row even if prod+style match
    for row in ledger_rows:
        prod = str(row.get('production', '')).strip()
        style = str(row.get('style', '')).strip().upper()
        if not prod or not style:
            continue
        # ETD differentiates split shipments (same PO+style, different arrival dates)
        etd_part = str(row.get('etd') or 'no-etd').strip()
        key = f"{prod}::{style}::{etd_part}"
        # If same key still occurs (identical prod+style+etd), append an index
        if key in current_keys:
            i = 2
            while f"{key}::v{i}" in current_keys:
                i += 1
            key = f"{key}::v{i}"
        current_keys.add(key)
        ledger_map[key] = row

    changes = {
        'new': 0, 'updated_units': 0, 'updated_etd': 0,
        'arrived': 0, 'arrived_confirmed': 0, 'arrived_denied': 0,
        'miss_incremented': 0, 'incoming': 0, 'in_stock': 0,
    }

    # STEP 1: Process ledger rows
    for key, row in ledger_map.items():
        style = str(row.get('style', '')).strip().upper()
        brand_code = str(row.get('brand', '')).strip().upper()
        platform_brand = BRAND_CODE_MAP.get(brand_code, brand_code)
        try:
            units = int(row.get('units', 0) or 0)
        except (ValueError, TypeError):
            units = 0
        etd = row.get('etd')

        if key in snapshot['orders']:
            ex = snapshot['orders'][key]
            ex['last_seen_on_ledger'] = now
            ex['consecutive_misses'] = 0
            if ex.get('status') == 'in_stock':
                ex['status'] = 'incoming'
                ex['fell_off_date'] = None
            if ex.get('units') != units:
                ex.setdefault('units_history', []).append({'date': now, 'old': ex['units'], 'new': units})
                ex['units'] = units
                changes['updated_units'] += 1
            if ex.get('etd') != etd:
                changes['updated_etd'] += 1
                ex['etd'] = etd
            changes['incoming'] += 1
        else:
            snapshot['orders'][key] = {
                'production': str(row.get('production', '')).strip(),
                'po_name': str(row.get('poName', '')).strip(),
                'style': style, 'units': units,
                'brand_code': brand_code, 'brand': platform_brand,
                'etd': etd, 'status': 'incoming',
                'first_seen': now, 'last_seen_on_ledger': now,
                'fell_off_date': None, 'consecutive_misses': 0,
                'units_history': [],
            }
            changes['new'] += 1

    # STEP 2: Check for disappearances
    for key, order in snapshot['orders'].items():
        if order.get('status') != 'incoming':
            changes['in_stock'] += 1
            continue
        if key not in current_keys:
            misses = order.get('consecutive_misses', 0) + 1
            order['consecutive_misses'] = misses

            if misses >= MISS_THRESHOLD:
                # Archive validation
                should_flip = True
                if archive_data and current_ats:
                    sku = order.get('style', '').upper()
                    old = archive_data.get(sku)
                    cur = next((i for i in current_ats if i['sku'] == sku), None)
                    if old and cur:
                        cur_incoming = cur.get('incoming', 0)
                        if old['incoming'] > 0 and cur_incoming > 0:
                            should_flip = False
                            changes['arrived_denied'] += 1
                            print(f"  ⚠️ HELD: {key} — archive says still incoming")
                        elif old['incoming'] > 0 and cur_incoming == 0:
                            changes['arrived_confirmed'] += 1

                if should_flip:
                    order['status'] = 'in_stock'
                    order['fell_off_date'] = now
                    changes['arrived'] += 1
                    print(f"  📦 ARRIVED: {key} — {order.get('brand')} {order.get('units', 0):,} units")
            else:
                changes['miss_incremented'] += 1

    save_snapshot(snapshot)
    return changes, snapshot


# ══════════════════════════════════════════════════
#  INVENTORY BUILDER
# ══════════════════════════════════════════════════

def build_inventory(snapshot, ats_items, shipped_ledger=None):
    """Build inventory in same shape as the main platform's /inventory endpoint.
    
    Factors in shipped_out from the shipped ledger so SKUs that have been
    fully sold through don't rebound when committed/allocated drop.
    """
    ats_lookup = {}
    for item in ats_items:
        ats_lookup[item['sku']] = item
    
    shipped_lookup = (shipped_ledger or {}).get('skus', {})

    sku_agg = {}
    for key, order in snapshot.get('orders', {}).items():
        sku = order.get('style', '').upper()
        if not sku:
            continue
        if sku not in sku_agg:
            sku_agg[sku] = {'brand': order.get('brand', ''), 'arrived': 0, 'incoming': 0}
        if order.get('status') == 'in_stock':
            sku_agg[sku]['arrived'] += order.get('units', 0)
        else:
            sku_agg[sku]['incoming'] += order.get('units', 0)

    inventory = []
    for sku, agg in sku_agg.items():
        ats = ats_lookup.get(sku, {})
        committed = ats.get('committed', 0)
        allocated = ats.get('allocated', 0)
        arrived_original = agg['arrived']
        incoming = agg['incoming']
        brand = agg['brand']
        
        # ── Shipped deduction ──
        # Pull how many units this SKU has permanently shipped out
        shipped_out = 0
        if sku in shipped_lookup:
            shipped_out = shipped_lookup[sku].get('shipped_out', 0)
        
        # Effective arrived = what came in minus what's already shipped
        # This prevents sold-out SKUs from rebounding when committed drops
        arrived_effective = max(0, arrived_original - shipped_out)

        inventory.append({
            'sku': sku,
            'brand': brand,
            'brand_abbr': brand,
            'brand_full': BRAND_FULL_NAMES.get(brand, brand),
            'name': f"{brand} {sku}",
            'jtw': arrived_effective, 'tr': 0, 'dcw': 0, 'qa': 0,
            'incoming': incoming,
            'committed': committed,
            'allocated': allocated,
            'total_ats': arrived_effective + incoming - abs(committed) - abs(allocated),
            'total_warehouse': arrived_effective,
            'container': '', 'receive_date': '', 'lot_number': '',
            'image': '',
            # Metadata for transparency
            '_arrived_original': arrived_original,
            '_shipped_out': shipped_out,
        })

    return inventory


# ══════════════════════════════════════════════════
#  SYNC ENGINE — Runs hourly
# ══════════════════════════════════════════════════

_cache = {'inventory': [], 'snapshot': None, 'last_sync': None, 'syncing': False}
_cache_lock = threading.Lock()

def do_sync():
    with _cache_lock:
        if _cache['syncing']:
            return None
        _cache['syncing'] = True

    start = time.time()
    try:
        print("\n══════════════════════════════════════")
        print("  PRODUCTION SYNC")
        print("══════════════════════════════════════\n")

        # 1. Style ledger
        print("  📋 Fetching style ledger...")
        ledger = fetch_style_ledger()
        if not ledger:
            print("  ⚠ No ledger data")
            return None

        # 2. ATS deductions
        print("  📊 Fetching ATS deductions...")
        ats = fetch_ats_deductions()

        # 3. Archive for validation
        archive_data = None
        archive_name = None
        try:
            files = refresh_archive_index()
            if files:
                target = datetime.utcnow() - timedelta(hours=ARCHIVE_COMPARE_HOURS)
                best = min(files, key=lambda f: abs((f['dt'] - target).total_seconds()))
                print(f"  📂 Fetching archive: {best['name']}...")
                archive_data = fetch_archive_ats(best['path'])
                archive_name = best['name']
        except Exception as e:
            print(f"  ⚠ Archive skipped: {e}")

        # 4. Diff
        print("  🔄 Running diff...")
        changes, snapshot = run_diff(ledger, archive_data, ats)

        # 5. Update shipped ledger (tracks cumulative fulfillment to prevent rebound)
        print("  📤 Updating shipped ledger...")
        prod_arrived_by_sku = {}
        for key, order in snapshot.get('orders', {}).items():
            if order.get('status') == 'in_stock':
                sku = order.get('style', '').upper()
                prod_arrived_by_sku[sku] = prod_arrived_by_sku.get(sku, 0) + order.get('units', 0)
        shipped_ledger = update_shipped_ledger(ats, prod_arrived_by_sku)

        # 6. Build inventory (factors in shipped_out)
        print("  🔨 Building inventory...")
        inventory = build_inventory(snapshot, ats, shipped_ledger)

        with _cache_lock:
            _cache['inventory'] = inventory
            _cache['snapshot'] = snapshot
            _cache['shipped_ledger'] = shipped_ledger
            _cache['last_sync'] = datetime.utcnow().isoformat() + 'Z'
            _cache['syncing'] = False

        elapsed = time.time() - start
        # Count fully-shipped SKUs for reporting
        fully_shipped = sum(1 for rec in shipped_ledger.get('skus', {}).values()
                           if rec.get('shipped_out', 0) >= rec.get('arrived_total', 1))
        print(f"\n  ✅ DONE ({elapsed:.1f}s) — {len(inventory)} SKUs")
        print(f"     New: {changes['new']} | Arrived: {changes['arrived']} | "
              f"Incoming: {changes['incoming']} | In stock: {changes['in_stock']}")
        print(f"     Shipped ledger: {len(shipped_ledger.get('skus', {}))} SKUs tracked, "
              f"{fully_shipped} fully shipped out")
        if archive_name:
            print(f"     Archive: {archive_name} (confirmed: {changes['arrived_confirmed']}, denied: {changes['arrived_denied']})")

        return changes

    except Exception as e:
        import traceback
        print(f"  ❌ Sync failed: {e}")
        traceback.print_exc()
        with _cache_lock:
            _cache['syncing'] = False
        return None


# ══════════════════════════════════════════════════
#  ROUTES
# ══════════════════════════════════════════════════

@app.route('/production-inventory', methods=['GET', 'OPTIONS'])
def production_inventory():
    if request.method == 'OPTIONS':
        return '', 204
    with _cache_lock:
        inv = list(_cache['inventory'])
        last = _cache['last_sync']
    if not inv:
        do_sync()
        with _cache_lock:
            inv = list(_cache['inventory'])
            last = _cache['last_sync']
    return jsonify({'status': 'ok', 'inventory': inv, 'item_count': len(inv), 'last_sync': last})


@app.route('/production-sync', methods=['GET', 'POST', 'OPTIONS'])
def production_sync():
    if request.method == 'OPTIONS':
        return '', 204
    changes = do_sync()
    with _cache_lock:
        inv = _cache['inventory']
        last = _cache['last_sync']
    return jsonify({
        'status': 'ok' if changes else 'error',
        'inventory': inv, 'item_count': len(inv),
        'last_sync': last, 'changes': changes,
    })


@app.route('/production-snapshot', methods=['GET', 'OPTIONS'])
def production_snapshot():
    if request.method == 'OPTIONS':
        return '', 204
    with _cache_lock:
        snap = _cache.get('snapshot')
    if not snap:
        snap = load_snapshot()
    orders = list(snap.get('orders', {}).values())
    incoming = [o for o in orders if o.get('status') == 'incoming']
    in_stock = [o for o in orders if o.get('status') == 'in_stock']
    return jsonify({
        'last_check': snap.get('last_check'),
        'total': len(orders),
        'incoming_count': len(incoming), 'incoming_units': sum(o.get('units', 0) for o in incoming),
        'in_stock_count': len(in_stock), 'in_stock_units': sum(o.get('units', 0) for o in in_stock),
        'orders': snap.get('orders', {}),
    })


@app.route('/production-status', methods=['GET', 'OPTIONS'])
def production_status():
    if request.method == 'OPTIONS':
        return '', 204
    with _cache_lock:
        return jsonify({
            'status': 'ok', 'last_sync': _cache['last_sync'],
            'syncing': _cache['syncing'], 'inventory_count': len(_cache['inventory']),
        })


@app.route('/production-archive-index', methods=['GET', 'OPTIONS'])
def archive_index():
    if request.method == 'OPTIONS':
        return '', 204
    files = refresh_archive_index()
    limit = int(request.args.get('limit', 50))
    return jsonify({
        'total': len(files), 'showing': min(limit, len(files)),
        'files': [{'name': f['name'], 'date': f['dt'].strftime('%Y-%m-%d')} for f in files[:limit]],
    })


@app.route('/production-force-arrive', methods=['POST', 'OPTIONS'])
def force_arrive():
    if request.method == 'OPTIONS':
        return '', 204
    data = request.get_json(force=True, silent=True) or {}
    production = data.get('production', '')
    style = data.get('style', '')
    if not production:
        return jsonify({'error': 'production required'}), 400
    snapshot = load_snapshot()
    updated = 0
    for key, order in snapshot['orders'].items():
        if order['production'] != production:
            continue
        if style and order['style'] != style.upper():
            continue
        if order['status'] == 'in_stock':
            continue
        order['status'] = 'in_stock'
        order['fell_off_date'] = datetime.utcnow().isoformat() + 'Z'
        updated += 1
    if updated:
        save_snapshot(snapshot)
        with _cache_lock:
            _cache['inventory'] = []
    return jsonify({'ok': True, 'updated': updated})


@app.route('/production-force-incoming', methods=['POST', 'OPTIONS'])
def force_incoming():
    if request.method == 'OPTIONS':
        return '', 204
    data = request.get_json(force=True, silent=True) or {}
    production = data.get('production', '')
    style = data.get('style', '')
    if not production:
        return jsonify({'error': 'production required'}), 400
    snapshot = load_snapshot()
    updated = 0
    for key, order in snapshot['orders'].items():
        if order['production'] != production:
            continue
        if style and order['style'] != style.upper():
            continue
        if order['status'] == 'incoming':
            continue
        order['status'] = 'incoming'
        order['fell_off_date'] = None
        order['consecutive_misses'] = 0
        updated += 1
    if updated:
        save_snapshot(snapshot)
        with _cache_lock:
            _cache['inventory'] = []
    return jsonify({'ok': True, 'updated': updated})


@app.route('/production-discrepancies', methods=['GET', 'OPTIONS'])
def production_discrepancies():
    """
    Compare production data vs warehouse data to find discrepancies.
    This is READ-ONLY reporting — warehouse numbers never affect the production inventory.
    
    Shows:
      - SKUs where production says it arrived but warehouse shows less (potential receiving issue)
      - SKUs where warehouse shows more than production recorded (unexplained stock)
      - SKUs with zero-out events (warehouse dropped to 0)
    """
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        # Get current snapshot
        with _cache_lock:
            snap = _cache.get('snapshot')
        if not snap:
            snap = load_snapshot()
        
        # Get current ATS (for warehouse comparison only)
        ats = fetch_ats_deductions()
        ats_lookup = {item['sku']: item for item in ats}
        
        # Also get warehouse totals from a fresh ATS fetch
        ats_bytes = _dbx_download(DROPBOX_ATS_PATH)
        warehouse_lookup = {}
        if ats_bytes:
            wb = openpyxl.load_workbook(BytesIO(ats_bytes), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                if not row[0]:
                    continue
                sku = str(row[0]).strip().upper()
                jtw = int(row[6] or 0) if len(row) > 6 else 0
                tr = int(row[7] or 0) if len(row) > 7 else 0
                dcw = int(row[8] or 0) if len(row) > 8 else 0
                qa = int(row[9] or 0) if len(row) > 9 else 0
                warehouse_lookup[sku] = jtw + tr + dcw + qa
            wb.close()
        
        # Aggregate production data per SKU
        prod_arrived = {}  # sku -> total arrived units from production
        prod_incoming = {}  # sku -> total still-incoming units
        for key, order in snap.get('orders', {}).items():
            sku = order.get('style', '').upper()
            if not sku:
                continue
            units = order.get('units', 0) or 0
            if order.get('status') == 'in_stock':
                prod_arrived[sku] = prod_arrived.get(sku, 0) + units
            else:
                prod_incoming[sku] = prod_incoming.get(sku, 0) + units
        
        discrepancies = []
        for sku, prod_total in prod_arrived.items():
            wh_total = warehouse_lookup.get(sku, 0)
            ats_item = ats_lookup.get(sku, {})
            committed = abs(ats_item.get('committed', 0))
            allocated = abs(ats_item.get('allocated', 0))
            diff = prod_total - wh_total
            
            # Flag if there's a material mismatch
            if abs(diff) >= 10:  # ignore rounding-level differences
                discrepancies.append({
                    'sku': sku,
                    'brand': ats_item.get('brand', ''),
                    'production_arrived': prod_total,
                    'warehouse_actual': wh_total,
                    'difference': diff,
                    'committed': committed,
                    'allocated': allocated,
                    'sold_through_expected': committed + allocated,
                    'notes': (
                        'Factory shipped more than warehouse received' if diff > 0 else
                        'Warehouse has more than production recorded'
                    )
                })
        
        # Sort by absolute difference (biggest first)
        discrepancies.sort(key=lambda d: abs(d['difference']), reverse=True)
        
        total_diff_units = sum(abs(d['difference']) for d in discrepancies)
        
        return jsonify({
            'status': 'ok',
            'count': len(discrepancies),
            'total_discrepancy_units': total_diff_units,
            'checked_skus': len(prod_arrived),
            'discrepancies': discrepancies,
            'note': 'Warehouse data used for comparison only — production inventory numbers are unchanged',
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/production-reset', methods=['POST', 'GET', 'OPTIONS'])
def production_reset():
    """Wipe the snapshot and re-sync from scratch. Use if data is corrupted.
    
    Query param ?wipe_shipped=true also wipes the shipped ledger (full reset).
    """
    if request.method == 'OPTIONS':
        return '', 204
    try:
        empty = {'last_check': None, 'orders': {}}
        save_snapshot(empty)
        
        wipe_shipped = request.args.get('wipe_shipped', '').lower() in ('true', '1', 'yes')
        if wipe_shipped:
            save_shipped_ledger({'last_update': None, 'skus': {}})
            print("🗑️ Shipped ledger also wiped")
        
        with _cache_lock:
            _cache['inventory'] = []
            _cache['snapshot'] = None
            _cache['shipped_ledger'] = None
            _cache['last_sync'] = None
        print("🗑️ Snapshot wiped — re-syncing from scratch...")
        changes = do_sync()
        with _cache_lock:
            count = len(_cache['inventory'])
            last = _cache['last_sync']
        return jsonify({
            'status': 'ok',
            'message': f'Snapshot wiped and re-synced{" (shipped ledger also wiped)" if wipe_shipped else ""}',
            'inventory_count': count,
            'last_sync': last,
            'changes': changes,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/production-shipped-ledger', methods=['GET', 'OPTIONS'])
def production_shipped_ledger():
    """View the shipped ledger — tracks cumulative shipments per SKU."""
    if request.method == 'OPTIONS':
        return '', 204
    try:
        with _cache_lock:
            ledger = _cache.get('shipped_ledger')
        if not ledger:
            ledger = load_shipped_ledger()
        
        # Summary stats
        skus = ledger.get('skus', {})
        total_shipped = sum(rec.get('shipped_out', 0) for rec in skus.values())
        fully_shipped = [sku for sku, rec in skus.items()
                        if rec.get('shipped_out', 0) >= rec.get('arrived_total', 1)]
        partially_shipped = [sku for sku, rec in skus.items()
                            if 0 < rec.get('shipped_out', 0) < rec.get('arrived_total', 1)]
        
        return jsonify({
            'last_update': ledger.get('last_update'),
            'tracked_skus': len(skus),
            'total_units_shipped': total_shipped,
            'fully_shipped_skus': len(fully_shipped),
            'partially_shipped_skus': len(partially_shipped),
            'skus': skus,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/production-shipped-adjust', methods=['POST', 'OPTIONS'])
def production_shipped_adjust():
    """Manually adjust shipped_out for a SKU. Use for corrections.
    Body: { sku: "...", shipped_out: 1000 }
    """
    if request.method == 'OPTIONS':
        return '', 204
    data = request.get_json(force=True, silent=True) or {}
    sku = data.get('sku', '').upper()
    new_shipped = int(data.get('shipped_out', 0))
    if not sku:
        return jsonify({'error': 'sku required'}), 400
    try:
        ledger = load_shipped_ledger()
        if sku not in ledger['skus']:
            ledger['skus'][sku] = {
                'shipped_out': 0, 'peak_committed': 0, 'peak_allocated': 0,
                'history': [], 'first_tracked': datetime.utcnow().isoformat() + 'Z',
            }
        prev = ledger['skus'][sku]['shipped_out']
        ledger['skus'][sku]['shipped_out'] = max(0, new_shipped)
        ledger['skus'][sku]['history'].append({
            'date': datetime.utcnow().isoformat() + 'Z',
            'manual_adjustment': True,
            'from': prev,
            'to': new_shipped,
        })
        save_shipped_ledger(ledger)
        with _cache_lock:
            _cache['inventory'] = []  # Force rebuild
        return jsonify({'ok': True, 'sku': sku, 'previous': prev, 'new': new_shipped})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════════
#  READ-ONLY PROXY — Main API data (overrides, banners, etc.)
#  
#  All proxy endpoints are GET-only. The production platform
#  inherits this data from the main platform but CANNOT modify it.
#  POST/PUT/DELETE requests get a clear "read-only" error.
# ══════════════════════════════════════════════════

# Block all write attempts to proxied endpoints
READONLY_ENDPOINTS = [
    '/overrides', '/banner-rules', '/prepack-defaults',
    '/suppression-overrides', '/deduction-assignments',
    '/manual-allocations', '/allocations', '/saved-catalogs',
]

@app.before_request
def block_writes_to_readonly():
    """Return a clear error if frontend tries to save/edit anything."""
    if request.method in ('POST', 'PUT', 'DELETE'):
        for endpoint in READONLY_ENDPOINTS:
            if request.path == endpoint or request.path.startswith(endpoint + '/'):
                return jsonify({
                    'error': 'read_only',
                    'message': 'Production platform is read-only. Make changes on the main inventory platform.'
                }), 403


# Proxy endpoints to main API (so frontend doesn't need to know about 2 servers)
@app.route('/production', methods=['GET', 'OPTIONS'])
def proxy_production():
    """Proxy to main API's /production endpoint for style ledger display data."""
    if request.method == 'OPTIONS':
        return '', 204
    try:
        resp = http_requests.get(f'{MAIN_API_URL}/production', timeout=20)
        return jsonify(resp.json())
    except:
        return jsonify({'production': []})


@app.route('/overrides', methods=['GET', 'OPTIONS'])
def proxy_overrides():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        resp = http_requests.get(f'{MAIN_API_URL}/overrides', timeout=10)
        return jsonify(resp.json())
    except:
        return jsonify({'overrides': {}})


@app.route('/allocations', methods=['GET', 'OPTIONS'])
def proxy_allocations():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        resp = http_requests.get(f'{MAIN_API_URL}/allocations', timeout=10)
        return jsonify(resp.json())
    except:
        return jsonify([])


@app.route('/manual-allocations', methods=['GET', 'OPTIONS'])
def proxy_manual_alloc():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        resp = http_requests.get(f'{MAIN_API_URL}/manual-allocations', timeout=10)
        return jsonify(resp.json())
    except:
        return jsonify([])


@app.route('/suppression-overrides', methods=['GET', 'OPTIONS'])
def proxy_suppression():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        resp = http_requests.get(f'{MAIN_API_URL}/suppression-overrides', timeout=10)
        return jsonify(resp.json())
    except:
        return jsonify({'overrides': []})


@app.route('/banner-rules', methods=['GET', 'OPTIONS'])
def proxy_banners():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        resp = http_requests.get(f'{MAIN_API_URL}/banner-rules', timeout=10)
        return jsonify(resp.json())
    except:
        return jsonify({'rules': []})


@app.route('/prepack-defaults', methods=['GET', 'OPTIONS'])
def proxy_prepack():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        resp = http_requests.get(f'{MAIN_API_URL}/prepack-defaults', timeout=10)
        return jsonify(resp.json())
    except:
        return jsonify({'defaults': {}})


@app.route('/deduction-assignments', methods=['GET', 'OPTIONS'])
def proxy_deductions():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        resp = http_requests.get(f'{MAIN_API_URL}/deduction-assignments', timeout=10)
        return jsonify(resp.json())
    except:
        return jsonify({})


@app.route('/apo', methods=['GET', 'OPTIONS'])
def proxy_apo():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        resp = http_requests.get(f'{MAIN_API_URL}/apo', timeout=10)
        return jsonify(resp.json())
    except:
        return jsonify({'apo': []})


@app.route('/dropbox-photos', methods=['GET', 'OPTIONS'])
def proxy_photos():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        resp = http_requests.get(f'{MAIN_API_URL}/dropbox-photos', timeout=10)
        return jsonify(resp.json())
    except:
        return jsonify({'codes': []})


@app.route('/image/<path:sku>', methods=['GET'])
def proxy_image(sku):
    """Proxy image requests to main API."""
    try:
        brand = request.args.get('brand', '')
        resp = http_requests.get(f'{MAIN_API_URL}/image/{sku}?brand={brand}', timeout=10, stream=True)
        return resp.content, resp.status_code, {'Content-Type': resp.headers.get('Content-Type', 'image/jpeg')}
    except:
        return '', 404


# ══════════════════════════════════════════════════
#  HEALTH CHECK — Root URL
# ══════════════════════════════════════════════════

@app.route('/', methods=['GET'])
def health():
    with _cache_lock:
        count = len(_cache['inventory'])
        last = _cache['last_sync']
    return jsonify({
        'status': 'ok',
        'service': 'Production Inventory API',
        'inventory_count': count,
        'last_sync': last,
    })


# ══════════════════════════════════════════════════
#  HOURLY CRON
# ══════════════════════════════════════════════════

def start_cron():
    def loop():
        time.sleep(8)
        print("🏭 Initial production sync...", flush=True)
        do_sync()
        while True:
            time.sleep(SYNC_INTERVAL)
            print("⏰ Hourly production sync...", flush=True)
            do_sync()
    t = threading.Thread(target=loop, daemon=True)
    t.start()

start_cron()


# ══════════════════════════════════════════════════
#  STARTUP
# ══════════════════════════════════════════════════

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    print(f"\n🏭 Production Inventory Server starting on port {port}")
    print(f"   Ledger: {DROPBOX_LEDGER_FOLDER}")
    print(f"   ATS:    {DROPBOX_ATS_PATH}")
    print(f"   Archive:{DROPBOX_ARCHIVE_PATH}")
    print(f"   S3:     {S3_BUCKET}/{S3_SNAPSHOT_KEY}")
    print(f"   Main API proxy: {MAIN_API_URL}\n")
    app.run(host='0.0.0.0', port=port, debug=False)
